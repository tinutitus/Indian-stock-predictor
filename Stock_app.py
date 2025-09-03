import yfinance as yf
import pandas as pd
import numpy as np
import datetime, requests, os, platform, subprocess
from io import BytesIO
from xgboost import XGBClassifier
from sklearn.model_selection import train_test_split
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule

# -----------------------------
# Step 1: Fetch Midcap + Smallcap tickers dynamically
# -----------------------------
HEADERS = {"User-Agent": "Mozilla/5.0", "Referer": "https://www.niftyindices.com/"}

def fetch_index_tickers(url):
    r = requests.get(url, headers=HEADERS, timeout=20)
    r.raise_for_status()
    df = pd.read_csv(BytesIO(r.content))
    sym_col = [c for c in df.columns if "Symbol" in c or "SYMBOL" in c][0]
    tickers = df[sym_col].astype(str).str.strip().str.upper() + ".NS"
    return tickers.tolist()

midcap_url = "https://www.niftyindices.com/IndexConstituent/ind_niftymidcap100list.csv"
smallcap_url = "https://www.niftyindices.com/IndexConstituent/ind_niftysmallcap100list.csv"

try:
    midcap_tickers = fetch_index_tickers(midcap_url)
    smallcap_tickers = fetch_index_tickers(smallcap_url)
except Exception as e:
    print("⚠️ Could not fetch live tickers, using fallback list:", e)
    midcap_tickers = ["ABB.NS","AUROPHARMA.NS","BANKBARODA.NS","CANBK.NS"]
    smallcap_tickers = ["ABFRL.NS","CESC.NS","FORTIS.NS","IRCTC.NS"]

tickers = midcap_tickers + smallcap_tickers
print(f"✅ Loaded {len(midcap_tickers)} Midcap and {len(smallcap_tickers)} Smallcap tickers")

# -----------------------------
# Step 2: Parameters
# -----------------------------
start = datetime.datetime.now() - datetime.timedelta(days=365*2)
end = datetime.datetime.now()
target_pct = 5   # Target return % in 1 month
prob_threshold = 0.6  # ML probability filter
price_limit = 1000    # Shortlist price limit

all_results = []

# -----------------------------
# Step 3: Process each ticker
# -----------------------------
for ticker in tickers:
    try:
        df = yf.download(ticker, start=start, end=end, progress=False)
        if df.empty:
            continue

        # Technical Indicators
        df["Return_5d"] = df["Adj Close"].pct_change(5)
        df["Return_20d"] = df["Adj Close"].pct_change(20)
        df["Volatility"] = df["Adj Close"].pct_change().rolling(20).std()

        # Moving Averages
        df["MA20"] = df["Adj Close"].rolling(20).mean()
        df["MA50"] = df["Adj Close"].rolling(50).mean()
        df["MA200"] = df["Adj Close"].rolling(200).mean()

        # MACD
        ema12 = df["Adj Close"].ewm(span=12, adjust=False).mean()
        ema26 = df["Adj Close"].ewm(span=26, adjust=False).mean()
        df["MACD"] = ema12 - ema26

        # Bollinger Band %B
        df["BB_up"] = df["MA20"] + 2*df["Adj Close"].rolling(20).std()
        df["BB_down"] = df["MA20"] - 2*df["Adj Close"].rolling(20).std()
        df["BB_pct"] = (df["Adj Close"] - df["BB_down"]) / (df["BB_up"] - df["BB_down"])

        # RSI
        delta = df["Adj Close"].diff()
        gain = delta.clip(lower=0).rolling(14).mean()
        loss = -delta.clip(upper=0).rolling(14).mean()
        rs = gain / (loss + 1e-9)
        df["RSI"] = 100 - (100 / (1 + rs))

        # Volume % change
        df["Vol_Change"] = df["Volume"].pct_change()

        # Target variable
        df["Target"] = (df["Adj Close"].shift(-20) / df["Adj Close"] - 1) * 100
        df["Target"] = (df["Target"] >= target_pct).astype(int)
        df.dropna(inplace=True)

        if len(df) < 60:
            continue

        # Features
        features = ["Return_5d","Return_20d","Volatility",
                    "MA20","MA50","MA200","MACD","BB_pct",
                    "RSI","Vol_Change"]
        X = df[features]
        y = df["Target"]

        # Train-test split
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, shuffle=False)

        # Model = XGBoost
        model = XGBClassifier(n_estimators=200, max_depth=5, learning_rate=0.05,
                              subsample=0.8, colsample_bytree=0.8,
                              use_label_encoder=False, eval_metric="logloss", random_state=42)
        model.fit(X_train, y_train)

        # Latest prediction
        latest_features = X.iloc[-1:].values
        prob_up = model.predict_proba(latest_features)[0][1]

        current_price = round(df["Adj Close"].iloc[-1], 2)

        # Expected Prices
        expected_price_1m = round(current_price * (1 + prob_up * (target_pct/100)), 2)
        expected_price_1y = round(current_price * ((1 + prob_up * (target_pct/100)) ** 12), 2)

        # Differences
        diff_1m = round(expected_price_1m - current_price, 2)
        diff_1y = round(expected_price_1y - current_price, 2)

        # Percentage Changes
        pctchange_1m = round((diff_1m / current_price) * 100, 2) if current_price else 0
        pctchange_1y = round((diff_1y / current_price) * 100, 2) if current_price else 0

        # Final Score (blend ML + Momentum20 + RSI_norm)
        rsi_norm = (df["RSI"].iloc[-1] / 100) if not np.isnan(df["RSI"].iloc[-1]) else 0.5
        momentum20 = df["Return_20d"].iloc[-1] if not np.isnan(df["Return_20d"].iloc[-1]) else 0
        final_score = (0.5 * prob_up) + (0.25 * momentum20) + (0.25 * rsi_norm)

        all_results.append({
            "Company": ticker,
            "Price": current_price,
            "ExpectedPrice_1M": expected_price_1m,
            "ExpectedPrice_1Y": expected_price_1y,
            "Diff_1M": diff_1m,
            "Diff_1Y": diff_1y,
            "PctChange_1M": pctchange_1m,
            "PctChange_1Y": pctchange_1y,
            "Momentum_5d": round(df["Return_5d"].iloc[-1], 3),
            "Momentum_20d": round(momentum20, 3),
            "Volatility": round(df["Volatility"].iloc[-1], 3),
            "RSI": round(df["RSI"].iloc[-1], 2),
            "MACD": round(df["MACD"].iloc[-1], 3),
            "BB_pct": round(df["BB_pct"].iloc[-1], 3),
            "Vol_Change": round(df["Vol_Change"].iloc[-1], 3),
            "ProbUp1M": round(prob_up, 3),
            "FinalScore": round(final_score, 3)
        })
    except Exception as e:
        print(f"Error processing {ticker}: {e}")

# -----------------------------
# Step 4: Output
# -----------------------------
pred_df = pd.DataFrame(all_results)

if not pred_df.empty:
    # Rank by FinalScore
    pred_df["Rank"] = pred_df["FinalScore"].rank(ascending=False, method="dense").astype(int)

    # Column order
    cols = ["Company","Price","ExpectedPrice_1M","ExpectedPrice_1Y",
            "Diff_1M","Diff_1Y","PctChange_1M","PctChange_1Y","FinalScore","Rank"] + \
           [c for c in pred_df.columns if c not in ["Company","Price","ExpectedPrice_1M","ExpectedPrice_1Y",
                                                    "Diff_1M","Diff_1Y","PctChange_1M","PctChange_1Y",
                                                    "FinalScore","Rank"]]
    pred_df = pred_df[cols]

    # Shortlist
    shortlist = pred_df[(pred_df["Price"] <= price_limit) & (pred_df["ProbUp1M"] >= prob_threshold)]
    shortlist = shortlist.sort_values(by="FinalScore", ascending=False)

    # Save CSV
    pred_df.to_csv("all_predictions.csv", index=False)
    shortlist.to_csv("shortlist.csv", index=False)

    # Save Excel with formatting
    excel_file = "predictions.xlsx"
    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        pred_df.to_excel(writer, sheet_name="All Predictions", index=False)
        shortlist.to_excel(writer, sheet_name="Shortlist", index=False)

    wb = load_workbook(excel_file)
    for sheet_name in ["All Predictions","Shortlist"]:
        ws = wb[sheet_name]

        # Conditional formatting for PctChange
        for col_letter in ["G","H"]:  # PctChange_1M, PctChange_1Y
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{ws.max_row}",
                ColorScaleRule(start_type="num", start_value=-50, start_color="FF0000",
                               mid_type="num", mid_value=0, mid_color="FFFFFF",
                               end_type="num", end_value=50, end_color="00FF00")
            )

        # Gradient for FinalScore
        ws.conditional_formatting.add(
            f"I2:I{ws.max_row}",
            ColorScaleRule(start_type="num", start_value=0, start_color="FF0000",
                           mid_type="num", mid_value=0.5, mid_color="FFFF00",
                           end_type="num", end_value=1, end_color="00FF00")
        )

    wb.save(excel_file)

    # Auto-open Excel file
    system = platform.system()
    try:
        if system == "Windows":
            os.startfile(excel_file)
        elif system == "Darwin":  # macOS
            subprocess.call(["open", excel_file])
        else:  # Linux
            subprocess.call(["xdg-open", excel_file])
    except Exception as e:
        print("⚠️ Could not auto-open Excel:", e)

    print("✅ Analysis complete!")
    print(f"📊 Excel with formatting saved as {excel_file} (auto-opened)")
else:
    print("No data processed.")
